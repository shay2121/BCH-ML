import random as random
import numpy as np
import csv
from openpyxl import load_workbook
from array import array

class Code:
    """A simple example class"""
    i = 12345
    def __init__(self, n, k):
      self.n=n
      self.k=k
      self.m=n-k
      self.G=self.readCsvG()
    def readCsvG(self):
      r = np.empty([self.k, self.n], dtype=int)
      wb = load_workbook('CFG/codes.xlsx')
      sheet = wb.get_sheet_by_name('bch' + str(self.k) + 'x' + str(self.n))
      for i in range(2, 2 + self.k - 1):
        for j in range(1, 1 + self.n ):
          r[i - 2, j-1] = sheet.cell(row=i, column=j).value.__int__()
      return r

    def flip(self,p):
      return 0 if random.random() < 0.5 else 1

    def generateCodeWord(self,snr,globalInd=-1):
      R = np.true_divide(self.k, self.n)
      globalSnrUp = snr + 10 * np.log10(2.0 * R)
      factor = 2 * pow(10.0, (globalSnrUp / 20.0))
      noise = np.random.normal(0, 1, self.n) / factor
      w = np.random.randint(2, size=self.k)
      x = np.remainder(np.dot(w, self.G), 2)
      y = np.float16(x + noise)
      y = ["%.4f" % xx for xx in y]
      if globalInd>0:
        answer = y + [x[globalInd]]
      else:
        answer=y
      row4Write = np.asarray(answer)
      return row4Write

    def generateData4Training(self,globalInd,snr,numOfWords, ind, dataFileName1, dataFileName2):
      step = numOfWords / 100
      dataFile1 = open(dataFileName1, "wb")
      dataFile2 = open(dataFileName2, "wb")
      global dataWriter1
      global dataWriter2
      dataWriter1 = csv.writer(dataFile1, delimiter=',', quotechar='"', quoting=csv.QUOTE_NONE)
      dataWriter2 = csv.writer(dataFile2, delimiter=',', quotechar='"', quoting=csv.QUOTE_NONE, )
      r1 = np.append(self.k, np.append(ind, np.zeros(self.n- 3)))
      r1 = np.append(self.n, r1)
      r1 = np.append(numOfWords, r1)
      r1=r1.astype(int)
      dataWriter1.writerow(r1)
      dataWriter2.writerow(r1)
      for i in range(1, numOfWords):
        r = self.generateCodeWord(snr,globalInd)
        dataWriter1.writerow(r)
        r = self.generateCodeWord(snr,globalInd)
        dataWriter2.writerow(r)
      dataFile1.close()
      dataFile2.close()


